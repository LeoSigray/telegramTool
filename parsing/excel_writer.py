"""
Запись результатов парсинга в Excel-файл.
Каждая категория — отдельный лист. Поддержка дедупликации и сортировки.
"""

import os
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .categories import sanitize_sheet_name

# Колонки для папок (полная информация с подписчиками)
FOLDER_COLUMNS = ["Название", "Username", "Ссылка", "Тип", "Подписчиков", "Город"]

# Колонки для ZIP (только ссылки, без подписчиков)
ZIP_COLUMNS = ["Ссылка", "Источник"]

# Общий набор колонок (используем максимальный)
ALL_COLUMNS = ["Название", "Username", "Ссылка", "Тип", "Подписчиков", "Город", "Источник"]

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_ALIGN = Alignment(horizontal="center")

COLUMN_WIDTHS = {
    "Название": 35,
    "Username": 22,
    "Ссылка": 40,
    "Тип": 12,
    "Подписчиков": 14,
    "Город": 15,
    "Источник": 20,
}


def _apply_header(ws):
    """Применяет стиль заголовка к первой строке."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _set_column_widths(ws):
    """Устанавливает ширину колонок."""
    for col_idx, cell in enumerate(ws[1], 1):
        col_letter = cell.column_letter
        width = COLUMN_WIDTHS.get(cell.value, 15)
        ws.column_dimensions[col_letter].width = width


def _get_or_create_sheet(wb: Workbook, name: str):
    """Получает или создает лист с заголовками."""
    sheet_name = sanitize_sheet_name(name)
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    ws.append(ALL_COLUMNS)
    _apply_header(ws)
    _set_column_widths(ws)
    return ws


def _read_existing_links(ws) -> set:
    """Читает существующие ссылки из листа для дедупликации."""
    links = set()
    link_col = None
    for col_idx, cell in enumerate(ws[1], 0):
        if cell.value == "Ссылка":
            link_col = col_idx
            break
    if link_col is None:
        return links
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > link_col and row[link_col]:
            links.add(str(row[link_col]).strip())
    return links


def _sort_sheet_by_subscribers(ws):
    """Сортирует данные листа по количеству подписчиков (убывание)."""
    sub_col = None
    for col_idx, cell in enumerate(ws[1], 0):
        if cell.value == "Подписчиков":
            sub_col = col_idx
            break
    if sub_col is None:
        return

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))

    if not rows:
        return

    def sort_key(r):
        try:
            val = r[sub_col] if len(r) > sub_col else 0
            return int(val) if val else 0
        except (ValueError, TypeError):
            return 0

    rows.sort(key=sort_key, reverse=True)

    # Перезаписываем данные (без заголовка)
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(ALL_COLUMNS) + 1):
            ws.cell(row=row_idx, column=col_idx, value=None)

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)


class ParsingExcelWriter:
    """
    Менеджер записи результатов парсинга в Excel.

    Использование:
        writer = ParsingExcelWriter("data/parsed_chats.xlsx")
        writer.add_folder_chats("Криптовалюта и P2P", [...])
        writer.add_zip_chats("Ставки", [...])
        writer.save()
    """

    def __init__(self, file_path: str):
        self.file_path = file_path
        if os.path.exists(file_path):
            self.wb = load_workbook(file_path)
        else:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            self.wb = Workbook()
            # Удаляем дефолтный лист
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]

    def add_folder_chats(self, category: str, chats: list[dict], sort_by_subs: bool = True):
        """
        Добавляет чаты из парсинга папок в лист категории.

        chats: список словарей с ключами:
            title, username, link, chat_type, subscribers, city
        """
        ws = _get_or_create_sheet(self.wb, category)
        existing = _read_existing_links(ws)
        added = 0

        for chat in chats:
            link = chat.get("link", "")
            if link in existing:
                continue
            existing.add(link)
            ws.append([
                chat.get("title", ""),
                chat.get("username", ""),
                link,
                chat.get("chat_type", ""),
                chat.get("subscribers", 0),
                chat.get("city", ""),
                "Папка TG",
            ])
            added += 1

        if sort_by_subs and added > 0:
            _sort_sheet_by_subscribers(ws)

        return added

    def add_zip_chats(self, category: str, links: list[str], source: str = "ZIP архив"):
        """
        Добавляет ссылки из ZIP-файла в лист категории.

        links: список ссылок (строки)
        source: описание источника (напр., "ZIP / Беларусь")
        """
        ws = _get_or_create_sheet(self.wb, category)
        existing = _read_existing_links(ws)
        added = 0

        for link in links:
            if link in existing:
                continue
            existing.add(link)
            ws.append([
                "",       # Название (неизвестно из ZIP)
                "",       # Username
                link,     # Ссылка
                "",       # Тип
                "",       # Подписчиков
                "",       # Город
                source,   # Источник
            ])
            added += 1

        return added

    def save(self):
        """Сохраняет Excel-файл."""
        # Удаляем пустой лист Sheet если вдруг остался
        if "Sheet" in self.wb.sheetnames and len(self.wb.sheetnames) > 1:
            del self.wb["Sheet"]
        self.wb.save(self.file_path)
        self.wb.close()

    def get_stats(self) -> dict[str, int]:
        """Возвращает статистику: {название_листа: кол-во_строк}."""
        stats = {}
        for name in self.wb.sheetnames:
            ws = self.wb[name]
            count = ws.max_row - 1  # минус заголовок
            if count > 0:
                stats[name] = count
        return stats
