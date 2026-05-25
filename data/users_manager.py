"""
Менеджер списка пользователей (лист 'Users' в targets.xlsx).

Функции:
  add_users_manually()   — добавить пользователей вручную (user_id / @username)
  parse_members_from_chat() — спарсить участников чата и добавить в Users
"""

import asyncio
import os

import openpyxl
from telethon.tl.functions.channels import GetParticipantsRequest
from telethon.tl.types import (
    ChannelParticipantsSearch,
    InputPeerChannel,
    InputPeerChat,
    PeerChannel,
    PeerChat,
    PeerUser,
)

from config import DATA_DIR, EXCEL_FILE

USERS_SHEET = "Users"
CHATS_SHEET = "Chats"


# ─────────────────────────── Excel helpers ───────────────────────────────────

def _ensure_workbook() -> openpyxl.Workbook:
    """Создаёт файл targets.xlsx с нужными листами если не существует."""
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = USERS_SHEET
        ws.append(["user_id", "username", "comment"])
        ws_c = wb.create_sheet(CHATS_SHEET)
        ws_c.append(["chat_id", "username", "comment"])
        wb.save(EXCEL_FILE)
        return wb

    return openpyxl.load_workbook(EXCEL_FILE)


def _get_or_create_sheet(wb: openpyxl.Workbook, name: str, headers: list) -> openpyxl.worksheet.worksheet.Worksheet:
    """Возвращает лист по имени, создаёт с заголовками если нет."""
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
        ws.append(headers)
    # Если лист пустой — добавь заголовки
    if ws.max_row == 0 or ws.cell(1, 1).value != headers[0]:
        ws.insert_rows(1)
        for col, h in enumerate(headers, 1):
            ws.cell(1, col, h)
    return ws


def _existing_users(ws) -> set:
    """
    Возвращает множество уже добавленных идентификаторов:
    строковые user_id и нижний регистр username (без @).
    """
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        uid = row[0] if len(row) > 0 else None
        uname = row[1] if len(row) > 1 else None
        if uid is not None:
            seen.add(str(uid))
        if uname:
            seen.add(str(uname).lstrip("@").lower())
    return seen


# ─────────────────────────── Добавление вручную ──────────────────────────────

def add_users_manually(entries: list) -> dict:
    """
    Добавляет пользователей в конец листа 'Users', пропускает дубли.

    entries — список dict {"user_id": ..., "username": ..., "comment": ...}
    Возвращает {"added": N, "skipped": N}
    """
    wb = _ensure_workbook()
    ws = _get_or_create_sheet(wb, USERS_SHEET, ["user_id", "username", "comment"])

    existing = _existing_users(ws)
    added = skipped = 0

    for e in entries:
        uid_raw = str(e.get("user_id", "")).strip()
        uname_raw = str(e.get("username", "")).strip().lstrip("@")
        comment = str(e.get("comment", "")).strip() or None

        # Проверка дубля
        uid_key = uid_raw if uid_raw else None
        uname_key = uname_raw.lower() if uname_raw else None

        if (uid_key and uid_key in existing) or (uname_key and uname_key in existing):
            skipped += 1
            continue

        row_uid = int(uid_raw) if uid_raw.lstrip("-").isdigit() else (None if not uid_raw else uid_raw)
        row_uname = f"@{uname_raw}" if uname_raw else None

        ws.append([row_uid, row_uname, comment])

        if uid_key:
            existing.add(uid_key)
        if uname_key:
            existing.add(uname_key)
        added += 1

    wb.save(EXCEL_FILE)
    return {"added": added, "skipped": skipped}


# ─────────────────────────── Парсинг участников чата ─────────────────────────

async def _fetch_all_members(client, entity, limit_per_request: int = 200) -> list:
    """
    Скачивает всех участников через GetParticipantsRequest (постранично).
    Работает только для супергрупп/каналов. Для обычных групп — через get_participants.
    """
    members = []

    # Для обычных групп (Chat, not Channel) — другой способ
    try:
        from telethon.tl.types import Chat
        if isinstance(entity, Chat):
            async for user in client.iter_participants(entity):
                members.append(user)
            return members
    except Exception:
        pass

    # Супергруппа / канал
    offset = 0
    filter_ = ChannelParticipantsSearch("")
    while True:
        result = await client(GetParticipantsRequest(
            channel=entity,
            filter=filter_,
            offset=offset,
            limit=limit_per_request,
            hash=0,
        ))
        if not result.users:
            break
        members.extend(result.users)
        offset += len(result.users)
        if offset >= result.count:
            break
        await asyncio.sleep(1)

    return members


async def parse_members_from_chat(
    client,
    chat_link: str,
    comment: str = "",
    progress_cb=None,
) -> dict:
    """
    Парсит участников чата/группы и добавляет их в лист 'Users'.

    chat_link — @username, t.me/username или числовой ID
    comment   — метка/комментарий для записей в Excel
    progress_cb(current, total) — опциональный колбэк прогресса

    Возвращает {"parsed": N, "added": N, "skipped": N, "error": str|None}
    """
    # Разбор ссылки → идентификатор
    link = chat_link.strip()
    if link.startswith("https://t.me/") or link.startswith("http://t.me/"):
        link = link.split("t.me/")[-1].split("/")[0]
    if link.startswith("@"):
        link = link[1:]

    # Попробуем как число
    entity_id = None
    try:
        entity_id = int(link)
    except ValueError:
        pass

    try:
        entity = await client.get_entity(entity_id if entity_id else link)
    except Exception as e:
        return {"parsed": 0, "added": 0, "skipped": 0, "error": str(e)}

    try:
        members = await _fetch_all_members(client, entity)
    except Exception:
        # Fallback: iter_participants (работает без прав)
        try:
            members = []
            async for user in client.iter_participants(entity):
                members.append(user)
        except Exception as e2:
            return {"parsed": 0, "added": 0, "skipped": 0, "error": str(e2)}

    total = len(members)
    entries = []
    for i, user in enumerate(members):
        if progress_cb:
            progress_cb(i + 1, total)

        # Пропускаем ботов
        if getattr(user, "bot", False):
            continue

        uid = user.id
        uname = user.username or ""
        entries.append({"user_id": uid, "username": uname, "comment": comment})

    stats = add_users_manually(entries)
    return {
        "parsed": total,
        "added": stats["added"],
        "skipped": stats["skipped"],
        "error": None,
    }
