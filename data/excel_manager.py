import os
import re
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import EXCEL_FILE

# --- Стили ---

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_HEADER_ALIGN = Alignment(horizontal="center")

# Системные листы — не трогаем при парсинге
_SYSTEM_SHEETS = {"Users", "Chats", "DM_Log", "Chat_Log", "Invite_Log"}

# Колонки для листов-категорий (парсинг)
PARSED_COLUMNS = ["Название", "Username", "Ссылка", "Тип", "Подписчиков", "Город", "Источник"]

_COLUMN_WIDTHS = {
    "Название": 35, "Username": 22, "Ссылка": 40, "Тип": 12,
    "Подписчиков": 14, "Город": 15, "Источник": 20,
    "username": 25, "user_id": 18, "chat_id": 20, "comment": 30,
    "timestamp": 20, "account": 20, "target": 25, "status": 15, "error": 35,
}


# ========================================================================
#  Внутренние хелперы
# ========================================================================

def _apply_header(ws, columns):
    ws.append(columns)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    for cell in ws[1]:
        width = _COLUMN_WIDTHS.get(cell.value, 15)
        ws.column_dimensions[cell.column_letter].width = width


def _sanitize_sheet_name(name: str) -> str:
    name = (name or "Другое").strip()
    name = re.sub(r"[\[\]\*:/\\\?\n]", " ", name)
    return name[:31]


def _ensure_excel():
    """Создает Excel файл с базовыми листами, если его нет."""
    if os.path.exists(EXCEL_FILE):
        return

    os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
    wb = Workbook()

    ws_users = wb.active
    ws_users.title = "Users"
    _apply_header(ws_users, ["username", "user_id", "comment"])
    ws_users.append(["durov", "", "пример по username"])
    ws_users.append(["", "123456789", "пример по user_id"])

    ws_chats = wb.create_sheet("Chats")
    _apply_header(ws_chats, ["username", "chat_id", "comment"])
    ws_chats.append(["telegram", "", "пример по username"])
    ws_chats.append(["", "-1001234567890", "пример по chat_id"])

    for sheet_name in ("DM_Log", "Chat_Log", "Invite_Log"):
        ws = wb.create_sheet(sheet_name)
        _apply_header(ws, ["timestamp", "account", "target", "status", "error"])

    wb.save(EXCEL_FILE)
    print(f"Создан {EXCEL_FILE} — заполни 'Users' и 'Chats'.")


def _open_workbook(read_only=False):
    _ensure_excel()
    return load_workbook(EXCEL_FILE, read_only=read_only)


def _normalize_username(val) -> str | None:
    if not val:
        return None
    s = str(val).strip()
    return s.lstrip("@") if s else None


def _col_idx(headers: list, names: list):
    for name in names:
        if name in headers:
            return headers.index(name)
    return None


# ========================================================================
#  Загрузка таргетов (Users / Chats)
# ========================================================================

def load_users() -> list[dict]:
    wb = _open_workbook(read_only=True)
    ws = wb["Users"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx_username = _col_idx(headers, ["username", "юзернейм", "user"])
    idx_uid = _col_idx(headers, ["user_id", "id", "userid"])

    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        username = _normalize_username(row[idx_username] if idx_username is not None else None)
        user_id = row[idx_uid] if idx_uid is not None else None
        if user_id:
            try:
                user_id = int(str(user_id).strip())
            except ValueError:
                user_id = None
        if username or user_id:
            users.append({"user_id": user_id, "username": username})
    wb.close()
    return users


def load_chats() -> list[dict]:
    wb = _open_workbook(read_only=True)
    ws = wb["Chats"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx_username = _col_idx(headers, ["username", "юзернейм", "chat"])
    idx_cid = _col_idx(headers, ["chat_id", "id", "chatid"])

    chats = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        username = _normalize_username(row[idx_username] if idx_username is not None else None)
        chat_id = row[idx_cid] if idx_cid is not None else None
        if chat_id:
            try:
                chat_id = int(str(chat_id).strip())
            except ValueError:
                chat_id = None
        if username or chat_id:
            chats.append({"chat_id": chat_id, "username": username})
    wb.close()
    return chats


# ========================================================================
#  Логирование
# ========================================================================

def log_action(sheet_name, account, target, status, error=""):
    """
    Записывает лог действия в соответствующий лист.
    Если файл занят другой программой — сохраняет во временный файл рядом.
    """
    wb = _open_workbook()
    ws = wb[sheet_name]
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        str(account), str(target), status,
        str(error) if error else "",
    ])

    # Пробуем сохранить напрямую
    try:
        wb.save(EXCEL_FILE)
        wb.close()
        return
    except PermissionError:
        pass

    # Файл занят (открыт в Excel/PyCharm) — сохраняем рядом как log_pending.xlsx
    wb.close()
    pending = os.path.join(os.path.dirname(EXCEL_FILE), "log_pending.xlsx")
    try:
        if os.path.exists(pending):
            # Объединяем с уже накопленным pending файлом
            existing = load_workbook(pending)
            if sheet_name in existing.sheetnames:
                ex_ws = existing[sheet_name]
            else:
                ex_ws = existing.create_sheet(sheet_name)
                ex_ws.append(["timestamp", "account", "target", "status", "error"])
            ex_ws.append([
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                str(account), str(target), status, str(error) if error else "",
            ])
            existing.save(pending)
            existing.close()
        else:
            wb2 = load_workbook(EXCEL_FILE)
            wb2[sheet_name].append([
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                str(account), str(target), status, str(error) if error else "",
            ])
            wb2.save(pending)
            wb2.close()
        print(f"  [лог] targets.xlsx занят — запись сохранена в log_pending.xlsx")
    except Exception as e:
        print(f"  [лог] Не удалось записать лог: {e}")


def log_dm(account, target, status, error=""):
    log_action("DM_Log", account, target, status, error)


def log_chat(account, target, status, error=""):
    log_action("Chat_Log", account, target, status, error)


def log_invite(account, target, status, error=""):
    log_action("Invite_Log", account, target, status, error)


# ========================================================================
#  Парсинг: запись чатов в листы-категории
# ========================================================================

def _get_or_create_category_sheet(wb, category: str):
    """Получает или создает лист-категорию с заголовками."""
    name = _sanitize_sheet_name(category)
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(title=name)
    _apply_header(ws, PARSED_COLUMNS)
    return ws


def _read_existing_links(ws) -> set:
    """Читает все ссылки из листа для дедупликации."""
    links = set()
    link_col = None
    for col_idx, cell in enumerate(ws[1], 0):
        if cell.value == "Ссылка":
            link_col = col_idx
            break
    if link_col is None:
        return links
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > link_col and row[link_col]:
            links.add(str(row[link_col]).strip())
    return links


def _sort_sheet_by_subscribers(ws):
    """Сортирует данные листа по подписчикам (убывание)."""
    sub_col = None
    for col_idx, cell in enumerate(ws[1], 0):
        if cell.value == "Подписчиков":
            sub_col = col_idx
            break
    if sub_col is None:
        return

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return

    def sort_key(r):
        try:
            val = r[sub_col] if len(r) > sub_col else 0
            return int(val) if val else 0
        except (ValueError, TypeError):
            return 0

    rows.sort(key=sort_key, reverse=True)

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(PARSED_COLUMNS) + 1):
            ws.cell(row=row_idx, column=col_idx, value=None)

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def add_folder_chats(category: str, chats: list[dict], sort_by_subs: bool = True) -> int:
    """
    Добавляет чаты из парсинга папок в лист-категорию.

    chats: [{"title", "username", "link", "chat_type", "subscribers", "city"}, ...]
    Возвращает кол-во добавленных.
    """
    wb = _open_workbook()
    ws = _get_or_create_category_sheet(wb, category)
    existing = _read_existing_links(ws)
    added = 0

    for chat in chats:
        link = chat.get("link", "")
        if link in existing:
            continue
        existing.add(link)
        ws.append([
            chat.get("title", ""),
            chat.get("username", ""),
            link,
            chat.get("chat_type", ""),
            chat.get("subscribers", 0),
            chat.get("city", ""),
            "Папка TG",
        ])
        added += 1

    if sort_by_subs and added > 0:
        _sort_sheet_by_subscribers(ws)

    wb.save(EXCEL_FILE)
    wb.close()
    return added


def add_zip_chats(category: str, links: list[str], source: str = "ZIP") -> int:
    """
    Добавляет ссылки из ZIP-архива в лист-категорию.
    Возвращает кол-во добавленных.
    """
    wb = _open_workbook()
    ws = _get_or_create_category_sheet(wb, category)
    existing = _read_existing_links(ws)
    added = 0

    for link in links:
        if link in existing:
            continue
        existing.add(link)
        ws.append(["", "", link, "", "", "", source])
        added += 1

    wb.save(EXCEL_FILE)
    wb.close()
    return added


def get_category_stats() -> dict[str, int]:
    """Возвращает {категория: кол-во чатов} для всех листов-категорий."""
    wb = _open_workbook(read_only=True)
    stats = {}
    for name in wb.sheetnames:
        if name in _SYSTEM_SHEETS:
            continue
        ws = wb[name]
        count = ws.max_row - 1
        if count > 0:
            stats[name] = count
    wb.close()
    return stats


def load_chats_by_category(categories: list[str]) -> list[dict]:
    """
    Загружает чаты из указанных листов-категорий.
    Возвращает list of dict с ключами: username, link, category.
    """
    wb = _open_workbook(read_only=True)
    chats = []
    seen = set()

    for cat_name in categories:
        if cat_name not in wb.sheetnames:
            continue
        ws = wb[cat_name]

        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        uname_idx = None
        link_idx = None
        for i, h in enumerate(headers):
            if h == "Username":
                uname_idx = i
            elif h == "Ссылка":
                link_idx = i

        for row in ws.iter_rows(min_row=2, values_only=True):
            username = None
            link = ""

            if uname_idx is not None and len(row) > uname_idx and row[uname_idx]:
                username = str(row[uname_idx]).strip().lstrip("@")
            if link_idx is not None and len(row) > link_idx and row[link_idx]:
                link = str(row[link_idx]).strip()

            # Если нет username — пробуем извлечь из ссылки
            if not username and link:
                username = _extract_username_from_link(link)

            if not username or username.lower() in seen:
                continue
            seen.add(username.lower())
            chats.append({"username": username, "link": link, "category": cat_name})

    wb.close()
    return chats


def get_category_names() -> list[str]:
    """Возвращает список названий категорий (листов)."""
    wb = _open_workbook(read_only=True)
    names = [n for n in wb.sheetnames if n not in _SYSTEM_SHEETS]
    wb.close()
    return names


# ========================================================================
#  Перенос чатов из категорий в лист Chats (для рассылки)
# ========================================================================

def _extract_username_from_link(link: str) -> str | None:
    """Извлекает username из ссылки https://t.me/username."""
    if not link:
        return None
    link = str(link).strip()
    # https://t.me/username или https://t.me/+invite
    if link.startswith("https://t.me/"):
        tail = link[len("https://t.me/"):]
        # Пропускаем приватные и invite-ссылки
        if tail.startswith("+") or tail.startswith("joinchat/"):
            return None
        # Убираем параметры
        tail = tail.split("?")[0].split("/")[0]
        return tail if tail else None
    return None


def export_categories_to_chats(categories: list[str] | None = None) -> int:
    """
    Переносит чаты из листов-категорий в лист 'Chats' для рассылки.
    Извлекает username из ссылки. Дедупликация по username.

    categories: список категорий для экспорта, или None = все.
    Возвращает кол-во добавленных.
    """
    wb = _open_workbook()
    ws_chats = wb["Chats"]

    # Определяем порядок колонок в Chats
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws_chats[1]]
    uname_col = _col_idx(headers, ["username", "юзернейм", "chat"])
    cid_col = _col_idx(headers, ["chat_id", "id", "chatid"])
    comment_col = _col_idx(headers, ["comment", "комментарий"])
    num_cols = len(headers)

    # Собираем существующие username
    existing = set()
    if uname_col is not None:
        for row in ws_chats.iter_rows(min_row=2, values_only=True):
            val = row[uname_col] if len(row) > uname_col else None
            if val:
                existing.add(str(val).strip().lower())

    if categories is None:
        categories = [n for n in wb.sheetnames if n not in _SYSTEM_SHEETS]

    added = 0
    for cat_name in categories:
        if cat_name not in wb.sheetnames:
            continue
        ws = wb[cat_name]

        cat_headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        src_uname_idx = None
        src_link_idx = None
        for i, h in enumerate(cat_headers):
            if h == "Username":
                src_uname_idx = i
            elif h == "Ссылка":
                src_link_idx = i

        for row in ws.iter_rows(min_row=2, values_only=True):
            username = None
            if src_uname_idx is not None and len(row) > src_uname_idx:
                username = _normalize_username(row[src_uname_idx])
            if not username and src_link_idx is not None and len(row) > src_link_idx:
                username = _extract_username_from_link(row[src_link_idx])
            if not username or username.lower() in existing:
                continue

            existing.add(username.lower())

            # Пишем в правильные колонки
            new_row = [""] * num_cols
            if uname_col is not None:
                new_row[uname_col] = username
            if comment_col is not None:
                new_row[comment_col] = cat_name
            ws_chats.append(new_row)
            added += 1

    wb.save(EXCEL_FILE)
    wb.close()
    return added


def import_from_parsed_excel(parsed_file: str) -> dict[str, int]:
    """
    Импортирует данные из отдельного parsed_chats.xlsx в targets.xlsx.
    Возвращает {категория: кол-во добавленных}.
    """
    if not os.path.exists(parsed_file):
        raise FileNotFoundError(f"Файл не найден: {parsed_file}")

    src_wb = load_workbook(parsed_file, read_only=True)
    stats = {}

    for sheet_name in src_wb.sheetnames:
        ws_src = src_wb[sheet_name]
        src_headers = [str(c.value).strip() if c.value else "" for c in next(ws_src.iter_rows(min_row=1, max_row=1))]

        # Проверяем что это лист-категория (есть колонка Ссылка)
        if "Ссылка" not in src_headers:
            continue

        # Находим индексы колонок в исходном файле
        idx_map = {h: i for i, h in enumerate(src_headers)}

        chats = []
        for row in ws_src.iter_rows(min_row=2, values_only=True):
            chat = {}
            for col_name in PARSED_COLUMNS:
                idx = idx_map.get(col_name)
                if idx is not None and len(row) > idx:
                    chat[col_name] = row[idx]
                else:
                    chat[col_name] = ""
            if chat.get("Ссылка"):
                chats.append(chat)

        if not chats:
            continue

        # Пишем в targets.xlsx
        wb = _open_workbook()
        ws = _get_or_create_category_sheet(wb, sheet_name)
        existing = _read_existing_links(ws)
        added = 0

        for chat in chats:
            link = str(chat.get("Ссылка", "")).strip()
            if link in existing:
                continue
            existing.add(link)
            ws.append([chat.get(c, "") for c in PARSED_COLUMNS])
            added += 1

        if added > 0:
            _sort_sheet_by_subscribers(ws)

        wb.save(EXCEL_FILE)
        wb.close()

        if added > 0:
            stats[sheet_name] = added

    src_wb.close()
    return stats
